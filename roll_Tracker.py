import time

from Functions import plc_communication, read_data, read_bool, read_int, Mtc_communication, check_exist
from datetime import *
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import shutil
import datetime
import snap7
import multiprocessing
import asyncio
import threading
import time
'''
* program to communicate with plc
* get data from DB and save it into excel sheet
* Author: Ahmed Taha
* Program Name : Pasaban Tracking system
* Version: V 1.0
Date : 31/07/2024
'''
# Define used param

# Plc parameter for line c , line B and line A
line_c_ip = '192.168.4.72'
line_b_ip = '192.168.4.71'
line_a_ip = '192.168.4.71'
#sheeter_b_ip = '192.168.4.170'
sheeter_a_ip = '192.168.4.70'
mtc_plc10 = '192.168.4.10'
rack = 0
slot = 2
# line_c_DB
robot_ream_B_DB = 221
robot_ream_A_DB = 220
robot_pallet_param = 401
# line B plc
layer_counter_DB = 4
ream_preselection_DB = 11
robot_param_DB = 4
ream_start_address = 10
reel_start_address = 2
previous_state = False
working_reel = "none"

# start address for Ream & Reel
reel_id_list = []
ream_id_list = []
pallet_id_dict = {}
reel_id_b_list = []
ream_id_b_list = []

# connecting to plc
plc_10 = Mtc_communication('192.168.4.10', 10)
#sheeter_b_plc = plc_communication(sheeter_b_ip, rack, 3)
sheeter_a_plc = plc_communication(sheeter_a_ip, rack, 3)
line_a_plc = plc_communication(line_a_ip,rack, slot)

# Design Excel Sheet
header_list = ["Reel ID", "Number Of sheets","Sheet To deliv_3", "rejected paper", "Data","Start_time" , "End_Time"]
sheet_name = 'rolls Data'

order_nr = (read_data(line_a_plc, datablock_address=51, start_address=34, data_size=10)).strip()
# print(file_day)
file_name = "database/roll_data.xlsx"
copy_file = "roll_data.xlsx"
print(file_name)
# file_name = "report.xlsx"

# Create Excel sheet
n = 0
if os.path.exists(file_name):
    work_book = load_workbook(file_name)
    work_sheet = work_book.active
    row = work_sheet.max_row
    n = row - 1
else:
    work_book = Workbook()
    work_sheet = work_book.active
    for i, header in enumerate(header_list):
        cell = work_sheet.cell(row=1, column=i + 1)
        cell.value = header
        cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick')
                             , bottom=Side(style='thick'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size='12', bold='Bold')
        cell.fill = PatternFill(fill_type='solid', fgColor="00339966")
        n = 0
        work_sheet.column_dimensions['B'].width = 30
        work_sheet.column_dimensions['C'].width = 30
        work_sheet.column_dimensions['A'].width = 30
        work_sheet.column_dimensions['D'].width = 30
        work_sheet.column_dimensions['E'].width = 30
        work_sheet.column_dimensions['F'].width = 30
        work_sheet.column_dimensions['G'].width = 30
        work_sheet.row_dimensions[1].height = 25
state = False
total_cut = 0
temp = 0
current_row = None
total_sheets =0
reel_id_unwind1 = read_data(sheeter_a_plc, 26, 2, 6)
reel_id_unwind2 = read_data(sheeter_a_plc, 26, 10, 6)
working_unwind = read_bool(sheeter_a_plc, 7, 0, 128, 65, 5)
reel_1_diameter = read_int(sheeter_a_plc, 16, 922, 2)
reel_2_diameter = read_int(sheeter_a_plc, 25, 42, 2)

if working_unwind == 0:
    working_reel = reel_id_unwind1
    current_row = check_exist(working_reel,file_name)
    print(f"working_reel = {reel_id_unwind1}")
    if current_row:
        total_sheets = work_sheet.cell(row=current_row, column=2).value
        n = current_row
    else:
        work_sheet.cell(row=n + 2, column=1).value = working_reel
        total_sheets = 0
        n = work_sheet.max_row +1
        #print(f"reel_1_diameter = {reel_1_diameter}")
elif working_unwind == 1:
    working_reel = reel_id_unwind2
    current_row = check_exist(working_reel, file_name)
    if current_row:
        total_sheets = work_sheet.cell(row=current_row, column= 2).value
        print(f"working_reel = {reel_id_unwind2}")
        n = current_row
    else:
        print(f"working_reel = {reel_id_unwind2}")
        work_sheet.cell(row=n + 2, column=1).value = working_reel
        #print(f"reel_2_diameter = {reel_2_diameter}")
        n = work_sheet.max_row +1
current_diameter = False
reel_diameter = 0
#def monitor(plc_ip, datablock_address, start_address, data_size, order_work_sheet , interval = 1):
last_value = None
previous_time = None
cutting_changes = []
total_sheets = 0
#start = time.time()
#print(start)
period = 0
while True:

    total_cut = read_int(sheeter_a_plc, 1100, 96, 2)
    splice_cycle = read_bool(sheeter_a_plc, 260, 0, 16, 2, 5)
    # print(f"total cut = {total_cut}")
    try:
        if (total_cut != last_value) or splice_cycle:
            if last_value is not None:
                #period = ((datetime.datetime.now() - previous_time).total_seconds()) / 60
                cutting_changes.append((last_value, period))
                print(f" Sheeter is running with speed {last_value} sheets/min, and duration  = {period} ")
                    # order_work_sheet.cell(row = n, column = 2).value = last_value
                    # order_work_sheet.cell(row = n, column = 3).value = period
                if splice_cycle and not previous_state:
                    n = work_sheet.max_row
                    for value, period in cutting_changes:
                        total_sheets += (value * period *.2)
                    total_sheets = round(total_sheets /60)
                    print(f"Total sheets  = {total_sheets} sheets")
                    work_sheet.cell(row= n, column=2).value = total_sheets
                    reel_id_unwind1 = read_data(sheeter_a_plc, 26, 2, 6)
                    reel_id_unwind2 = read_data(sheeter_a_plc, 26, 10, 6)
                    working_unwind = read_bool(sheeter_a_plc, 7, 0, 128, 65, 5)
                    if working_unwind == 0:
                        working_reel = reel_id_unwind1
                        #current_row = check_exist(working_reel, file_name)
                        print(f"working_reel = {reel_id_unwind1}")
                        work_sheet.cell(row = n+1, column=1).value = working_reel
                        total_sheets = 0
                        """
                        if current_row:
                            total_sheets = work_sheet.cell(row=current_row, column=2).value
                        else:
                            work_sheet.cell(row=n + 1, column=1).value = working_reel
                            total_sheets = 0
                            # print(f"reel_1_diameter = {reel_1_diameter}")
                        """
                    elif working_unwind == 1:
                        working_reel = reel_id_unwind2
                        print(f"working_reel = {reel_id_unwind2}")
                        total_sheets = 0
                        work_sheet.cell(row=n + 1, column=1).value = working_reel
                        """
                        current_row = check_exist(working_reel, file_name)
                        if current_row:
                            total_sheets = work_sheet.cell(row=current_row, column=2).value
                            print(f"working_reel = {reel_id_unwind2}")
                        else:
                            print(f"working_reel = {reel_id_unwind2}")
                            work_sheet.cell(row=n + 1, column=1).value = working_reel
                            total_sheets = 0
                        """
                    cutting_changes.clear()
                previous_state = splice_cycle
            #previous_time = datetime.datetime.now()
            last_value = total_cut
            period = 0
        period += 1
        print(f"period = {period}")
        time.sleep(.1)

        try:
            work_book.save(file_name)
            shutil.copy(file_name, copy_file)
            work_book.close()

        except:
            continue

    except Exception as e:
        print(f"error code : {e}")






