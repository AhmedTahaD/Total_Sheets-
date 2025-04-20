import datetime
import time
import snap7
from pyModbusTCP.client import ModbusClient
from openpyxl import *
import asyncio


def Mtc_communication(host, port):
    try:
        client = ModbusClient(host, port)
        client = ModbusClient(host="localhost", auto_open=True, auto_close=False)
        print(f"{host} is connected")

    except:
        print("enter a vaild ip")

# plc communication function
def plc_communication(ip, rack_num, slot_num):
    try:
        client = snap7.client.Client()
        client.connect(ip, rack_num, slot_num)
        print(f"{ip} is connected")
        return client
    except ConnectionError:
        print("Enter a valid ip")


# fetching data type string from datablock Function
def read_data(plc_ip, datablock_address, start_address, data_size):
    """read data from plc

    :param plc_ip: the client it should communicate with
    :param datablock_address: the datablock number
    :param start_address: the start address it should read from
    :param data_size: how many bits it should read
    :return: data that datablock contained
    """
    data = plc_ip.read_area(snap7.type.Areas.DB, datablock_address, start_address, data_size)
    data = bytearray(data)
    data = data.decode('utf-8')
    return data


# Read integer Value
def read_int(plc_ip, datablock_address, start_address, data_size):
    """
    read integer value
    :param plc_ip: the client it should communicate with
    :param datablock_address: the datablock number
    :param start_address: the start address it should read from
    :param data_size: how many bits it should read
    :return:
    """
    int_data = plc_ip.read_area(snap7.type.Areas.DB, datablock_address, start_address, data_size)
    int_data = int.from_bytes(int_data, byteorder='big', signed=False)
    return int_data


# read a boolean value from datablock
def read_bool(plc_ip, datablock_address, start_address, data_size, byte_num, bit_num):
    """
    read a digital value
    :param plc_ip: the client it should communicate with
    :param datablock_address: the datablock number
    :param start_address: the start address it should read from
    :param data_size: how many bits it should read
    :param byte_num: the data byte
    :param bit_num: bit num
    :return: the bit value
    """
    value = plc_ip.db_read(datablock_address, start_address, data_size)
    value = snap7.util.get_bool(value, byte_num, bit_num)
    return value
# check order name is exist
def check_exist(order_name, order_file):
    work_book = load_workbook(order_file)
    work_sheet = work_book.active
    for row in work_sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if order_name == cell.value:
                return cell.row
            """
            else :
                return False
            """
def get_orders_data(current_row, order_nr, orders_work_sheet):
    if current_row:
        total_sheets = orders_work_sheet.cell(row=current_row, column=2).value
        print(f"total_sheets = {total_sheets}")
        print(f"current_row_inside if = {current_row}")
    elif current_row is None:
        orders_row = orders_work_sheet.max_row
        current_row = orders_row + 1
        order_cell = orders_work_sheet.cell(row=current_row, column=1)
        order_cell.value = order_nr
        # current_row += 1
        total_sheets_cell = orders_work_sheet.cell(row=current_row, column=2)
        total_sheets_cell.value = 0
        print(f"current_row_inside else = {current_row}")
def monitor(plc_ip, datablock_address, start_address, data_size, order_work_sheet , interval = 1):
    last_value = None
    previous_time = None
    cutting_changes = []
    n = 2
    total_sheets = 0
    while True:
        total_cut = read_int(plc_ip, datablock_address, start_address, data_size)
        splice_cycle = read_bool(plc_ip, 260, 0, 16, 2, 5)
        #print(f"total cut = {total_cut}")
        print("Hi from monitor function")
        if total_cut != last_value or splice_cycle:
            if last_value != None  :
                period = ((datetime.datetime.now()- previous_time).total_seconds())/60
                cutting_changes.append((last_value, period))
                print(f" Sheeter is running with speed {last_value} sheets/min, and duration  = {period} ")
                    #order_work_sheet.cell(row = n, column = 2).value = last_value
                    #order_work_sheet.cell(row = n, column = 3).value = period
                if splice_cycle:
                    for value, period in  cutting_changes:
                        total_sheets += (value * period)

                    cutting_changes.clear()
                n+=1
            previous_time = datetime.datetime.now()
            last_value = total_cut






